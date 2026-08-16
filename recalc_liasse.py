#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
recalc_liasse.py — Recalcule les formules d'un classeur .xlsx via LibreOffice.

openpyxl ecrit des formules mais ne les evalue pas : a l'ouverture, les cellules
calculees apparaissent vides tant qu'Excel n'a pas recalcule. Ce script force le
calcul en pilotant LibreOffice en mode headless, puis reecrit le fichier avec les
valeurs en cache.

Autonome : ne depend que de LibreOffice (soffice) et de la bibliotheque standard.

Usage :
    python recalc_liasse.py fichier.xlsx [timeout_secondes]

Codes de retour :
    0  recalcul effectue
    1  echec du recalcul (le fichier d'origine est intact)
    2  LibreOffice introuvable
"""
import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path

MACRO = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script"
               script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''



def trouver_soffice():
    """Chemin de l'executable LibreOffice, ou None."""
    for nom in ('soffice', 'libreoffice'):
        chemin = shutil.which(nom)
        if chemin:
            return chemin
    for candidat in (
            '/usr/bin/soffice', '/usr/bin/libreoffice',
            '/usr/lib/libreoffice/program/soffice',
            '/opt/libreoffice/program/soffice',
            '/Applications/LibreOffice.app/Contents/MacOS/soffice'):
        if Path(candidat).exists():
            return candidat
    return None


def _installer_macro(soffice, profil, timeout=60):
    """Prepare un profil LibreOffice temporaire et y depose la macro Basic.

    Le profil doit etre initialise par LibreOffice lui-meme : c'est lui qui cree
    l'arborescence user/basic/Standard et les fichiers de declaration de
    bibliotheque. On se contente ensuite d'y ajouter le module.
    """
    url = Path(profil).as_uri()
    subprocess.run(
        [soffice, '--headless', '--terminate_after_init',
         '-env:UserInstallation=' + url],
        capture_output=True, timeout=timeout)

    dossier = Path(profil) / 'user' / 'basic' / 'Standard'
    if not dossier.exists():
        raise RuntimeError('LibreOffice n a pas cree de profil utilisable.')
    (dossier / 'Module1.xba').write_text(MACRO, encoding='utf-8')
    return url


def recalculer(chemin_xlsx, timeout=180, verbeux=True):
    """Recalcule le classeur sur place. Retourne True si le calcul a eu lieu."""
    chemin = Path(chemin_xlsx).resolve()
    if not chemin.exists():
        if verbeux:
            print('Fichier introuvable : ' + str(chemin), file=sys.stderr)
        return False

    soffice = trouver_soffice()
    if not soffice:
        if verbeux:
            print('LibreOffice introuvable (soffice absent du PATH). '
                  'Le classeur reste utilisable : ses formules seront '
                  'calculees a la premiere ouverture dans Excel.',
                  file=sys.stderr)
        return False

    sauvegarde = chemin.with_suffix(chemin.suffix + '.bak')
    shutil.copy2(chemin, sauvegarde)

    profil = tempfile.mkdtemp(prefix='lo_profil_')
    try:
        url = _installer_macro(soffice, profil, timeout=min(60, timeout))

        commande = [
            soffice, '--headless', '--norestore',
            '-env:UserInstallation=' + url,
            'vnd.sun.star.script:Standard.Module1.RecalculateAndSave'
            '?language=Basic&location=application',
            str(chemin),
        ]
        if shutil.which('timeout'):
            commande = ['timeout', str(timeout)] + commande

        t0 = time.time()
        resultat = subprocess.run(commande, capture_output=True,
                                  timeout=timeout + 15)
        duree = time.time() - t0

        if resultat.returncode != 0:
            if verbeux:
                print('LibreOffice a echoue (code ' + str(resultat.returncode) + ') : '
                      + resultat.stderr.decode('utf-8', 'ignore')[:400], file=sys.stderr)
            shutil.copy2(sauvegarde, chemin)
            return False

        if verbeux:
            print('Recalcul effectue en ' + format(duree, '.1f') + 's : ' + chemin.name)
        return True

    except subprocess.TimeoutExpired:
        if verbeux:
            print('Recalcul interrompu apres ' + str(timeout) + 's.', file=sys.stderr)
        shutil.copy2(sauvegarde, chemin)
        return False
    except Exception as e:
        if verbeux:
            print('Erreur de recalcul : ' + str(e), file=sys.stderr)
        shutil.copy2(sauvegarde, chemin)
        return False
    finally:
        shutil.rmtree(profil, ignore_errors=True)
        sauvegarde.unlink(missing_ok=True)


def controler(chemin_xlsx):
    """Compte les formules et les cellules d'erreur apres recalcul."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    ERREURS = ('#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!')
    wb_f = load_workbook(chemin_xlsx)
    wb_v = load_workbook(chemin_xlsx, data_only=True)
    formules = calculees = erreurs = 0
    detail = []
    for ws in wb_f.worksheets:
        for ligne in ws.iter_rows():
            for c in ligne:
                if isinstance(c.value, str) and c.value.startswith('='):
                    formules += 1
                    v = wb_v[ws.title][c.coordinate].value
                    if v is not None:
                        calculees += 1
                    if isinstance(v, str) and v in ERREURS:
                        erreurs += 1
                        if len(detail) < 20:
                            detail.append(ws.title + '!' + c.coordinate + ' = ' + v)
    return {'formules': formules, 'formules_calculees': calculees,
            'erreurs': erreurs, 'detail_erreurs': detail}


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    chemin = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 180

    if not trouver_soffice():
        return 2

    ok = recalculer(chemin, timeout)
    rapport = controler(chemin)
    if rapport:
        print('Formules            : ' + str(rapport['formules']))
        print('Formules calculees  : ' + str(rapport['formules_calculees']))
        print('Cellules en erreur  : ' + str(rapport['erreurs']))
        for d in rapport['detail_erreurs']:
            print('   ' + d)
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
